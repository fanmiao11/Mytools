#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import Workbook

def create_excel_file():
    """创建新的Excel文件"""
    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "工作表1"

        # 添加示例数据
        ws['A1'] = '姓名'
        ws['B1'] = '年龄'
        ws['C1'] = '部门'

        ws['A2'] = '张三'
        ws['B2'] = 25
        ws['C2'] = '技术部'

        ws['A3'] = '李四'
        ws['B3'] = 30
        ws['C3'] = '市场部'

        ws['A4'] = '王五'
        ws['B4'] = 28
        ws['C4'] = '人事部'

        # 保存文件
        filename = '新Excel文件.xlsx'
        wb.save(filename)
        print(f'Excel文件 "{filename}" 创建成功！')
        return True

    except ImportError:
        print('错误：需要安装 openpyxl 库')
        print('请运行：pip install openpyxl')
        return False
    except Exception as e:
        print(f'创建Excel文件时出错：{e}')
        return False

if __name__ == "__main__":
    create_excel_file()